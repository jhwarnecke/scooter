from django.shortcuts import render
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


# Create your views here.


def index(request):
    # reading excel sheet, index_col = none so we can get indices instead of names as headers
    # that's why we will always start with row '1' and not '0'
    # for columns we will start with '2' because there the values (variables) start
    global df
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    # get shape of excel sheet (to know number of iterations)
    global x
    x = df.shape

    # get Names of Models and Names of Cost-Structure-Elements into two Lists
    mylist = []
    myseclist = []
    
    # Hier wird die Liste mit der Kombi aus Namen des Anbieters und dem Namen des Modells befüllt
    for a in range(0, x[0]-1):
        mylist.append((df[0][a + 1])+" "+(df[1][a + 1]))
    # an dieser Stelle wird ähnlich wie oben eine Liste befüllt, aber hier mit den Namen der verschiedenen Spalten
    # gestartet wird hier erst bei "2", weil davor Anbieter und Modell liegen
    for b in range(2, x[1]):
        myseclist.append(df[b][0])
    print(myseclist)
    global dx
    dx = df.to_html(header=False, index=False)

    return render(request, "edittable/edittable.html", {"mylist": mylist, "myseclist": myseclist, "metable":dx})


def editfunc(request):
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    x = df.shape

    # variablen, die inputs speichert
    tarif = int(request.POST['modelname'])
    mycolumn = int(request.POST['column'])
    newvalue = float(request.POST['newvalue'])
    # Additionen um auf die richtige Zelle zuzugreifen in der xl
    tarif = tarif + 2
    mycolumn = mycolumn + 3
    # tabelle ändern
    workbook_name = 'scooter.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active
    page.cell(row = tarif, column = mycolumn).value = newvalue
    wb.save(filename=workbook_name)
    
    
    # Log Excel Sheet for input data
    # gleicher Aufbau wie für unseren anderen Log
    workbook_name2 = 'log_in.xlsx'
    wb2 = load_workbook(workbook_name2)
    page2 = wb2['edit']

    # New data to write:
    # hier nehmen wir wieder einmal die Zeit, dann den Anbieter mit angehängtem Modell (df[][],
    # als drittes den Spalten Namen aus der tabelle) und dann der neue Wert, der eingesetzt wird
    # hier wieder - 1 bei Tarif weil wir hier wieder von 0 zählen (wir sind wieder in python (df), nicht in der xl)
    new_companies2 = [datetime.now(), (df[0][tarif - 1])+" "+(df[1][tarif - 1]), df[mycolumn - 1][0], newvalue]

    page2.append(new_companies2)
    wb2.save(filename=workbook_name2)

    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    dx = df.to_html(header=False, index=False)

    return render(request, "edittable/ergebnis.html", {"metable":dx})


def addfuncstart(request):
    
    return render(request, "edittable/addfare.html", {"metable":dx})

def addfunc(request):
    # variablen, die inputs speichert
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)

    company = (request.POST['company'])
    model = (request.POST['model'])
    fix = float(request.POST['fix'])
    ppminute = float(request.POST['ppminute'])
    ppp = float(request.POST['ppp'])
    ppmonth = float(request.POST['ppmonth'])
    ppday = float(request.POST['ppday'])
    pcontingent = float(request.POST['pcontingent'])
    paftercontingent = float(request.POST['paftercontingent'])
    limit = float(request.POST['limit'])
    pafterlimit = float(request.POST['pafterlimit'])

    # tabelle ändern
    workbook_name = 'scooter.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active
    new_fare = [company,  model, fix, ppminute, ppp, ppmonth, ppday, pcontingent, paftercontingent, limit, pafterlimit]
    page.append(new_fare)
    wb.save(filename=workbook_name)

    # Log Excel Sheet for input data
    # gleicher Aufbau wie für unseren anderen Log
    workbook_name2 = 'log_in.xlsx'
    wb2 = load_workbook(workbook_name2)
    page2 = wb2['add']

    # New data to write:
    # hier nehmen wir wieder einmal die Zeit, dann den Anbieter mit angehängtem Modell (df[][],
    # als drittes den Spalten Namen aus der tabelle) und dann der neue Wert, der eingesetzt wird
    # hier wieder - 1 bei Tarif weil wir hier wieder von 0 zählen (wir sind wieder in python (df), nicht in der xl)
    
    new_companies2 = [datetime.now(), company,  model, fix, ppminute, ppp, ppmonth, ppday, pcontingent, paftercontingent, limit, pafterlimit]

    page2.append(new_companies2)
    wb2.save(filename=workbook_name2)

    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    dx = df.to_html(header=False, index=False)

    return render(request, "edittable/ergebnis.html", {"metable":dx})

def deletefuncstart(request):
    # get Names of Models and Names of Cost-Structure-Elements into two Lists
    mylist = []
    x = df.shape

    # Hier wird die Liste mit der Kombi aus Namen des Anbieters und dem Namen des Modells befüllt
    for a in range(0, x[0]-1):
        mylist.append((df[0][a + 1])+" "+(df[1][a + 1]))
    


    return render(request, "edittable/deletefare.html", {"mylist": mylist, "metable":dx})


def deletefunc(request):
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    row = int(request.POST['modelname'])
    # plus zwei, weil excel ab 1 anfängt und die Überschriften noch übersprungen werden müssen
    row = row + 2
    workbook_name = 'scooter.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active
    page.delete_rows(row)

    wb.save(filename=workbook_name)

    workbook_name2 = 'log_in.xlsx'
    wb2 = load_workbook(workbook_name2)
    page2 = wb2['delete']

    

    # New data to write:
    # hier nehmen wir wieder einmal die Zeit, dann den Anbieter mit angehängtem Modell (df[][],
    # als drittes den Spalten Namen aus der tabelle) und dann der neue Wert, der eingesetzt wird
    # hier wieder - 1 bei Tarif weil wir hier wieder von 0 zählen (wir sind wieder in python (df), nicht in der xl)
    new_companies2 = [datetime.now(), (df[0][row - 1])+" "+(df[1][row - 1])]

    page2.append(new_companies2)
    wb2.save(filename=workbook_name2)

    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    dx = df.to_html(header=False, index=False)

    return render(request, "edittable/ergebnis.html", {"metable":dx})


# test