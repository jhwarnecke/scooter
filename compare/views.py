from abc import ABC
from os import error, name
from django.shortcuts import render
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import base64
from matplotlib import pylab
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook







def index(request):
    # reading excel sheet, index_col = none so we can get indices instead of names as headers
    # that's why we will always start with row '1' and not '0'
    # for columns we will start with '2' because there the values (variables) start
    global df
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    # get shape of excel sheet (to know number of iterations)
    global x
    x = df.shape
    
    # Transformieren unserer Tabelle in html, um sie auf der Startseite zu zeigen
    dx = df.to_html(header=False, index=False)
    return render(request, "compare/input.html", {"metable":dx})


def calculation(request):
    df = pd.read_excel(r'scooter.xlsx', index_col=None, header=None)
    x = df.shape

    # Definitionen der Tage, für die spätere Abfrage nach Anzahl an Tagen
    global monday
    monday = False
    global tuesday
    tuesday = False
    global wednesday
    wednesday = False
    global thursday
    thursday = False
    global friday
    friday = False
    global saturday
    saturday = False
    global sunday
    sunday = False
    global monday2
    monday2 = False
    global tuesday2
    tuesday2 = False
    global wednesday2
    wednesday2 = False
    global thursday2
    thursday2 = False
    global friday2
    friday2 = False
    global saturday2
    saturday2 = False
    global sunday2
    sunday2 = False
    
    # reasons sind unsere Picklists aus denen der User auswählen kann,
    # wofür die jeweilige Strecke genutzt wird (technisch sonst kein Nutzen außer für log Datei)
    reason = request.POST["reason"]
    reason2 = request.POST["reason2"]
    
    
    ######## ToDo: Auskommentieren bzw löschen
    print(reason)
    print(reason2)

    # Abfrage der eingegebenen Werte für die Checkboxes der Tage, an denen gefahren wird
    day = request.POST.getlist("day")
    day2 = request.POST.getlist("day2")

    # Die Anzahl an Tagen muss größer null sein, weil sonst keine Berechnungen möglich sind
    # sollten keine Tage eingetragen sein, wird auf die Error Seite weitergeleitet
    if len(day) == 0:
        return(render(request, "error.html"))

    
    for z in range(0, len(day)):
        if day[z] == "monday":
            monday = True
        if day[z] == "tuesday":
            tuesday = True
        if day[z] == "wednesday":
            wednesday = True
        if day[z] == "thursday":
            thursday = True
        if day[z] == "friday":
            friday = True
        if day[z] == "saturday":
            saturday = True
        if day[z] == "sunday":
            sunday = True

    if day2 is not None:
        for z in range(0, len(day2)):
            if day2[z] == "monday":
                monday2 = True
            if day2[z] == "tuesday":
                tuesday2 = True
            if day2[z] == "wednesday":
                wednesday2 = True
            if day2[z] == "thursday":
                thursday2 = True
            if day2[z] == "friday":
                friday2 = True
            if day2[z] == "saturday":
                saturday2 = True
            if day2[z] == "sunday":
                sunday2 = True

    # Abfrage der Länge der Liste, um zu ermitteln, wieviele Tage ein Nutzer den Roller benutzen will (in der Woche Mo-So)
    numdays = len(day)
    numdays2 = len(day2)

    # Abfrage für die Zeit pro Nutzung und Abfrage für die Anzahl an Benutzungen pro Tag
    # ToDO: weg -> time_of_use = sum(time_dict.values())
    time_per_use = int(request.POST['usetime'])
    uses_per_day = int(request.POST['usesday'])

    # Kalkulation der gesamten Benutzungen aus der Anzahl an gefahrenen Tagen * Fahrten pro Tag * 4 (um die Monatszahlen zu bekommen)
    uses = numdays * uses_per_day * 4
    # GEsamte Nutzungsdauer kalkuliert aus der Länge einer Benutzung * der Benutzungen
    total_time = time_per_use * uses
    
    # Das gleiche wie oben, aber mit der zweiten Pendlerstrecke
    time_per_use2 = int(request.POST['usetime2'])
    uses_per_day2 = int(request.POST['usesday2'])
    uses2 = numdays2 * uses_per_day2 * 4
    total_time2 = time_per_use2 * uses2
    
    # Wenn an gleichen Tagen gefahren wird, werden dises von den Tagen pro use abgezogen 
    samedays = 0
    for z in range(0, len(day)):
        for y in range(0, len(day2)):
            if day[z] == day2[y]:
                samedays += 1

    # Addition der beiden Pendlerstrecken
    days_sum = (numdays + numdays2 - samedays) * 4
    uses_sum = uses + uses2
    total_time_sum = total_time + total_time2

    # Neuer Table um die Kosten und die dazugehörigen Namen der Modelle zu speichern
    mydf = pd.DataFrame()

    # iterations for all the models, save costs and name in df
    for i in range(0, x[1]):
        # normal wird unten definiert. Dazu sind zur Eingabe die Gesamten Tage, Benutzungen und Zeiten der beiden Nutzungen einzugeben
        # außerdem wird iteriert über i+1, weil die erste Zeile aus Überschriften besteht
        total_costs = round(normal(i, days_sum, uses_sum, total_time_sum, time_per_use, time_per_use2, 0), 2)
        # Hier wird der Name des Modells zwischengespeichert
        mytempname = ((df[0][i + 1])+" "+(df[1][i + 1]))
        # Und hier der errechnete Wert zusammen mit dem Namen des Modells gespeichert
        mytempdf = pd.DataFrame({'Name': [mytempname], 'Kosten': [total_costs]})
        # Die gespeicherten Werte werden dann in den vor der Schleife definierten Table geschrieben
        mydf = mydf.append(mytempdf, ignore_index=True)
        if df[9][i+1] != 0: # wenn es eine Zeitbeschränkung gibt und diese überschritten wird, wird ein zusätzlicher Tarif erstellt,
                            # "ohne Abstellen", der anzeigt wie hoch die Kosten sind, wenn man den Roller nicht zw.durch abstellt
            if df[9][i+1] < time_per_use or df[9][i+1] < time_per_use2:
                total_costs = round(normal(i, days_sum, uses_sum, total_time_sum, time_per_use, time_per_use2, 1), 2)
                mytempname = ((df[0][i + 1])+" "+(df[1][i + 1])+" ohne Abstellen")
                mytempdf = pd.DataFrame({'Name': [mytempname], 'Kosten': [total_costs]})
                mydf = mydf.append(mytempdf, ignore_index=True)
    
    # Heraus kommt eine Tabelle mit den Gesamtkosten und dem jeweiligen Modell
    # dann wird im folgenden dit Tabelle nach den Kosten sortiert und die Indizes resettet
    mydf = mydf.sort_values(by=['Kosten'])
    mydf = mydf.reset_index(drop=True)

    # Kosten und Name des günstigsten Modells in Variablen speichern
    minval = mydf.at[0,'Kosten']
    anbieter = mydf.at[0,'Name']

    # der Table mit den Ergebnissen wird zu einem html table umgeformt und die Indizes hinzugefügt
    myhtmldf = mydf.to_html(index = True)
    
    # Histogram Werte für die Ergebnis Seite
    # "height" ist dabei die Höhe der kosten und "bar_names" die Indizes, damit es einfacher zu lesen
    # ist und mit der Tabelle verglichen werden kann
    height = mydf['Kosten']
    bar_names = mydf.index
    
    # Definition folgt weiter unten, die Funktion nimmt die Kosten und den Index und baut daraus einen plot,
    # bzw macht aus dem plot ein .png, was dann auf der "Result page" angezeigt werden kann
    chart = get_plot(height, bar_names)
    
    # variable anzeige erstellen, bei True wird ein Zusatz in result.html ausgegeben
    if df[9][i+1] < time_per_use or df[9][i+1] < time_per_use2:
        anzeige = True
    else: anzeige = False
    
    # Loggen der Eingabe und Ausgabe Parameter in einer externen Excel Tabelle
    # Name der Datei wird eingelesen
    workbook_name = 'log_out.xlsx'
    # lesen und laden der Datei
    wb = load_workbook(workbook_name)
    # Aktivieren des Sheets
    page = wb.active

    # Daten, die im Log angezeigt werden sind die genaue Zeit der Eingabe, des Grundes der jeweiligen Strecke,
    # alle Eingaben (Fahrten pro Tag, Fahrtdauer im durchschnitt, Tage der Benutzung), sowie die Ergebnisse (Kosten + Modell)
    # New data to write:
    new_companies = [datetime.now(), reason, time_per_use, uses_per_day,
                    numdays, reason2, time_per_use2, uses_per_day2, numdays2, minval, anbieter]
    # hinzufügen zum excel sheet, als unterste Zeile
    page.append(new_companies)
    wb.save(filename=workbook_name)
    # Falls zwei gleichgünstig sind, wird auch der zweite gespeichert, dies wird unten in Zeile 228ff gemacht
    


    # Ausgabe auf der "Result" Seite
    # Falls 2 Tarife mit gleichen Kosten beide Ausgeben
    # Es werden die Werte für die minimalen Kosten (minval), den jeweiligen Namen des Anbieters,
    # wahlweise die Namen für ein Modell mit gleichen Kosten, dann der Table mit den sortierten
    # Werten und dem Bild des Histograms übergeben                            
    if mydf.at[0,'Kosten'] == mydf.at[1,'Kosten']:
        anbieter2 = mydf.at[1,'Name']
        # loggen des zweiten Modells
        workbook_name = 'log_out.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
        new_companies2 = [datetime.now(), reason, time_per_use, uses_per_day,
                    numdays, reason2, time_per_use2, uses_per_day2, numdays2, minval, anbieter2]
        page.append(new_companies2)
        wb.save(filename=workbook_name)
        return render(request, "compare/result.html", { "costs": minval, "name": anbieter,
                            "name2": anbieter2, "mytable": myhtmldf, "anzeige": anzeige, "chart": chart})
    # Standard Ausgabe, wenn es keine gleichen minimalen kosten gibt
    return render(request, "compare/result.html", { "costs": minval, "name": anbieter,
                            "mytable": myhtmldf, "name2": None, "anzeige": anzeige, "chart": chart})

                            
# definition of the cost function
# Die zuvor beschriebene "normal" Funktion bekommt die beschriebenen Werte + einen Wert
# für "extra_loop", wenn es Zeitbeschränkungen gibt und diese überschritten werden                            
def normal(j, days, uses, total_time, time_per_use, time_per_use2, extra_loop):
    # Grundpreis also die Grundgebühr pro Freischaltung * der Anzahl an Benutzungen                           
    grund = df[2][j+1] * uses
    # Preis pro Minute * der gesamten Nutzungszeit                            
    preispmin = df[3][j+1] * total_time
    # Preis für das Paket (wenn es eins ist)                            
    preisppak = df[4][j+1]
    # Monatlicher Preis                            
    preispmon = df[5][j+1]
    # Tageskarte bzw Preis dafür * Anzahl der Tage, an denen es genutzt werden würde                            
    preisptag = df[6][j+1] * days                            
    beschraenkung = 0
    if extra_loop == 1:
        beschraenkung = time_ones(j+1, time_per_use, time_per_use2)
    # Berechnung der Kontingent Preise in weiterer Funktion (übernimmt die Gesamtzeit der Nutzung und die Anzahl)                            
    konti = kontingent(j + 1, total_time, uses)
                                        #####ToDo: raus?                            
                                        #mylist = kontingent_loop(j+1, total_time)
                                        #konti_loop = mylist[0]
                                        #anteil_rest_konsten_konti_loop = mylist[1]

                                        #konti_plus = konti_loop - anteil_rest_konsten_konti_loop
    konti_plus = 0 # Zeile löschen, wenn die anderen Zeilen wieder reingenommen werden                       
    # Berechnung der anfallenden Kosten, die dann zurückgegeben werden                            
    costs = grund + preispmin + preisptag + preisppak + preispmon + konti_plus + beschraenkung + konti
    return costs

# def für normal von kontingent, keine Freischaltkosten nach Kontingent, Abo ist monatlich
def kontingent(j, total_time, uses):
    # if statement überprüft, ob es ein Kontingent ist, nur dann hat die Zeile einen Wert
    # größer 0 und gleichzeitig muss die total time überschritten sein                            
    if df[7][j] != 0 and df[7][j] < total_time:
        konti = (total_time - df[7][j]) * df[8][j]
    else:
        konti = 0
    return konti

#calculating extra costs, when time is higher then the per use time variable
def time_ones (j, time_per_use, time_per_use2):
        extra = 0
        # checks for input in field "Beschränkung"
        if df[9][j] != 0:
            # hier wird jeweils gecheckt, ob die einzelnen Pendlerstrecken, bzw deren Zeit pro Fahrt
            # größer sind als die unter Beschränkung eingetragenen Werte in der Tabelle (dem Excel Sheet)
            # wenn ja, dann wird die übrige Zeit aus der Differenz zwischen angegebener Zeit und Zeit aus
            # der Tabelle gegeben und mit dem jeweiligen Preis multipliziert, außerem *4 um es wieder auf 
            # 4 Wochen zu beziehen
            if df[9][j] < time_per_use:
                extra = (time_per_use - df[9][j]) * df[10][j] * 4
            # same thing für Pendlerstrecke 2
            if df[9][j] < time_per_use2:
                extra = extra + (time_per_use2 - df[9][j]) * df[10][j] * 4
        return extra

                            
# Definition für das Histogram:
# zuerste wird das Format der Figure festgelegt und decoded
# zuletzt muss das Bild wieder gelöscht werden (plt.clf()), damit 
# diese sich nicht überlagern                            
def get_graph():
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    graph = base64.b64encode(image_png)
    graph = graph.decode('utf-8')
    buffer.close()
    plt.clf()
    return graph

# Ausgabe für das eigentliche Histogram mit Werten für die Höhe der Kosten (height) und Namen bzw Indizes in diesem Fall
def get_plot(height, bar_names):
    y_pos = np.arange(len(bar_names))
    plt.bar(y_pos, height)
    plt.xticks(y_pos, bar_names)
    graph = get_graph()
    return graph


                                    # Die Funktion berechnet den Preis für einen Tarif mit Kontingent, welches (automatisch)
                                    # erneuerbar ist und dessen Minuten nicht verfallen!
                                    # def kontingent_loop(j, total_time):
                                    #     if df[7][j] != 0:
                                    #         konti =  0
                                    #         loop = 1
                                    #         new_total_time = total_time
                                    #         while loop == 1:
                                    #             new_total_time = new_total_time - df[7][j]
                                    #             if 0 < new_total_time:
                                    #                 loop = 1
                                    #                 konti = konti + df[4][j]

                                    #             else:
                                    #                 loop = 0
                                    #                 rest = - new_total_time
                                    #                 anteil_rest = rest / df[7][j]
                                    #                 anteil_rest_kosten = anteil_rest * df[4][j]
                                    #                 mylist = [konti, anteil_rest_kosten]

                                    #     else:
                                    #             konti = 0
                                    #             anteil_rest_kosten = 0
                                    #             mylist = [konti, anteil_rest_kosten]

                                    #     return mylist
